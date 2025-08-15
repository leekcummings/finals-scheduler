# ---------------------------------------------------------------------------- #
#                 Simmons University Finals Schedule Generator                 #
#                                 Lee Cummings                                 #
# ---------------------------------------------------------------------------- #

import numpy as np
import pandas as pd

# Max exams in one finals day
MAX_TESTS = 4

def cleanDF(df, courses):
    df = df.dropna()
    # Labs match the below pattern with a Course code, 3 numbers, and an L
    pattern = r'\w* \d{3}L-.*'
    df = df[~df['CourseSection'].str.contains(pattern)]
    df = df[~df['CourseSection'].str.contains('SIM 101')]
    # Remove classes that don't need final exams
    filter = [key for key, _ in courses.items() if courses[key]]
    df['CourseName'] = df['CourseSection'].str.extract(r'(\w* \d{3})')
    df = df[df['CourseName'].isin(filter)]
    
    return df

def createTimeslotGroups(df):
    df = df.drop('SID', axis=1).drop_duplicates().groupby(['Time Slot'], as_index=False)
    times = {}
    # Dictionary of timeslot per course
    for time, frame in df:
        for index, course in frame.iterrows():
            # KEY: Course ID
            # VALUE: Timeslot
            times[course['CourseSection']] = time[0]
    return times

def createStudentGroups(df, pop):
    courses = {}
    # Dictionary of students per course
    for course in pop:
        # Get a df of all students in a class
        students = df.loc[df['CourseSection'] == course]['SID']
        # KEY: Course ID
        # VALUE: Set of student
        courses[course] = set(students.to_numpy().flatten())
    return courses

def checkStudentConflicts(students, tests, i):
    for s in students:
        # If student has more than 2 tests in a day
        if s in tests[i // MAX_TESTS] and tests[i // MAX_TESTS][s] == 2:
            # Conflict found
            return True 
    # No conflicts
    return False 

def checkCourseTiming(times, schedule, c, i):
    currentTime = times[c]
    # For every course already in timeslot
    for otherCourse in schedule[i]:
        previousTime = times[otherCourse]
        # If times aren't the same
        if currentTime != previousTime:
            return True
    # All the same time
    return False

def checkRepeatedStudents(students, schedule, c, i):
    # Keep track of all student sets and number of students
    scheduledStudents = []
    totalStudents = 0
    # Students from unscheduled class
    scheduledStudents.append(students[c])
    totalStudents += len(students[c])

    for otherCourse in schedule[i]:
        # Students from previously scheduled classes
        scheduledStudents.append(students[otherCourse])
        totalStudents += len(students[otherCourse])
    
    # If the union of all student sets has repeats, 
    # it will be smaller than the total number of students
    if len(set.union(*scheduledStudents)) < totalStudents:
        # Repeats found
        return True
    # No repeats found
    return False

def updateStudentTests(students, tests, i):
    for s in students:
        tests[i // MAX_TESTS][s] = tests[i // MAX_TESTS].get(s, 0) + 1
    return tests


from openpyxl import Workbook
from openpyxl.styles import *





def export_excel(schedule, ntimes, path, compact=False, maxRows=15):
    wb = Workbook()
    sheet = wb.active

    cols = 3
    if compact: # Only change # cols in compact mode
        for slot in schedule:
            need = (len(slot)//maxRows)*2+3
            if need > cols:
                cols = need
    
    # Border definitions
    thin = Side(border_style="thin")
    thick = Side(border_style="thick")

    # Style for colunmn labels
    header = NamedStyle(name='header')
    header.font = Font(bold=True, size=14)
    header.alignment = Alignment(horizontal='center', vertical='center')
    header.border = Border(bottom=thick)

    for c in range(1,cols+3):
        sheet.cell(row=1, column=c).style = header
    for c in sheet[1]:
        sheet.column_dimensions[c.column_letter].width = 20

    # Labels
    sheet['A1'] = 'Exam Time'
    for c in range(2, cols+1): # Loop for compact columns
        if c%2 == 0:
            sheet.cell(row=1, column=c).value='Course'
        else:
            sheet.cell(row=1, column=c).value='Room'

    # Date and time inputs to be referenced in-sheet later
    sheet.cell(row=1, column=cols+2).value = 'Dates'
    sheet.cell(row=1, column=cols+2).font = Font(bold=True)
    for n in range(1, len(schedule)//ntimes+2):
        sheet.cell(row=n+1,column=cols+2).value = f"Day {n}"
        sheet.cell(row=n+1, column=cols+2).number_format = numbers.FORMAT_TEXT
    
    tRow = len(schedule)//ntimes+4
    sheet.cell(row=tRow, column=cols+2).value = 'Times'
    sheet.cell(row=tRow, column=cols+2).style = header
    sheet.cell(row=tRow, column=cols+2).font = Font(bold=True)
    for n in range(1, ntimes+1):
        sheet.cell(row=tRow+n, column=cols+2).value = f"Slot {n}"
        sheet.cell(row=tRow+n, column=cols+2).number_format = numbers.FORMAT_TEXT
    
    # Helper variables for formatting and references
    dataCol = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[cols+1] #stupidest implementation possible but it works lmao
    courseBorder = Border(bottom=thin, left=thin, right=thin, top=thin)
    roomBorder = Border(bottom=thin, top=thin, left=thin, right=thick)
    darkColor = PatternFill("solid",fgColor="FFa5c9ef")
    lightColor = PatternFill("solid",fgColor="FFd2e4f7") # NOTE: Maybe change colors to be more 'simmons'

    curRow = 2
    for n in range(len(schedule)):
        if len(schedule[n]) != 0:
            # Reference time and date
            sheet.cell(row=curRow, column=1).value = f'=_xlfn.TEXTJOIN(_xlfn.CHAR(10), TRUE, {dataCol}{n//ntimes+2}, {dataCol}{n%ntimes+tRow+1})'
            sheet.cell(row=curRow, column=1).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            if n%2==0:
                sheet.cell(row=curRow, column=1).fill = darkColor
            else:
                sheet.cell(row=curRow, column=1).fill = lightColor

            # Loop helpers for compact mode
            startRow = curRow
            maxRow = curRow+maxRows-1 

            curCol = 2

            sorted = schedule[n].copy()
            sorted.sort()
            for course in sorted:
                top = False # Used for border styling later
                if compact and curRow > maxRow: # Only change column in compact mode
                    curRow = startRow
                    curCol += 2
                    top = True
                
                # Set course and style 
                sheet.cell(row=curRow,column=curCol).value = course
                sheet.cell(row=curRow,column=curCol).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=curRow,column=curCol).border = courseBorder
                sheet.cell(row=curRow,column=curCol+1).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=curRow,column=curCol+1).border = roomBorder
                if n%2==0:
                    sheet.cell(row=curRow,column=curCol).fill = darkColor
                    sheet.cell(row=curRow,column=curCol+1).fill = darkColor
                else:
                    sheet.cell(row=curRow,column=curCol).fill = lightColor
                    sheet.cell(row=curRow,column=curCol+1).fill = lightColor
                
                # Ensure thick borders in compact mode
                if top:
                    sheet.cell(row=curRow,column=curCol).border = Border(bottom=thin, left=thin, right=thin, top=thick)
                    sheet.cell(row=curRow,column=curCol+1).border = Border(bottom=thin, left=thin, right=thick, top=thick)
                
                curRow+=1
            
            # Create thick border beneath current section, merge section label column
            if curCol == 2:
                for c in range(1, 3):
                    sheet.cell(row=curRow-1,column=c).border = Border(bottom=thick, left=thin, right=thin, top=thin)
                sheet.cell(row=curRow-1, column=3).border = Border(bottom=thick, left=thin, right=thick, top=thin)
                sheet.merge_cells(f"A{startRow}:A{curRow-1}")
            else:
                for c in range(1, curCol):
                    if c > 1 and c%2==1:
                        sheet.cell(row=maxRow,column=c).border = Border(bottom=thick, left=thin, right=thick, top=thin)
                    else:
                        sheet.cell(row=maxRow,column=c).border = Border(bottom=thick, left=thin, right=thin, top=thin)
                if curRow == startRow+1:
                    sheet.cell(row=curRow-1,column=curCol).border = Border(bottom=thick, left=thin, right=thin, top=thick)
                    sheet.cell(row=curRow-1,column=curCol+1).border = Border(bottom=thick, left=thin, right=thick, top=thick)
                else:
                    sheet.cell(row=curRow-1,column=curCol).border = Border(bottom=thick, left=thin, right=thin, top=thin)
                    sheet.cell(row=curRow-1,column=curCol+1).border = Border(bottom=thick, left=thin, right=thick, top=thin)
                sheet.merge_cells(f"A{startRow}:A{maxRow}")
                curRow = maxRow + 1

    wb.save(filename=path)

def generationStart(path, courses, maxTests, maxDays):
    # Import CSV of SID and Courses
    df = cleanDF(pd.read_csv(path), courses)
    courseTimes = createTimeslotGroups(df)

    # Array of most popular courses (most students)
    popularCourses = (df.groupby(['CourseSection'], sort=False)
                      .agg(NumStudents=('SID', 'count'))
                      .sort_values('NumStudents', ascending=False)
                      .index.to_numpy())
    courseStudent = createStudentGroups(df, popularCourses)

    # An array of exams, each index represents 1 timeslot
    schedule = []
    # To keep track of how many exams students have per day
    studentTests = {}

    for course in popularCourses:
        # The starting index for the schedule
        # Starting at 0 to ensure students get the earliest possible slot
        index = 0
        added = False
        students = courseStudent[course]

        while not added:
            # If the current day doesn't exist in studentTests
            if index // MAX_TESTS not in studentTests:
                studentTests[index // MAX_TESTS] = {}

            ### CASE 0: INDEX DOES NOT EXIST
            if index + 1 > len(schedule):
                studentConflict = checkStudentConflicts(students, studentTests, index)
                if not studentConflict:
                    # Append course in list, as it is first item in index
                    schedule.append([course])
                    studentTests = updateStudentTests(students, studentTests, index)
                    added = True
                else:
                    # Append a new empty list, to skip over index
                    schedule.append([])
            
            ### CASE 1: INDEX IS EMPTY
            elif len(schedule[index]) == 0:
                studentConflict = checkStudentConflicts(students, studentTests, index)
                if not studentConflict:
                    # Append course in list, as it is first item in index
                    schedule[index].append(course)
                    studentTests = updateStudentTests(students, studentTests, index)
                    added = True

            ### CASE 2: INDEX NOT EMPTY
            # If there are items in current index
            elif len(schedule[index]) > 0:
                differentTime = checkCourseTiming(courseTimes, schedule, course, index)
                studentConflict = checkStudentConflicts(students, studentTests, index)
                # If all times are same or no conflicts, add to schedule
                ### CASE 2A: CLASS TIMES ARE THE SAME
                if not differentTime and not studentConflict:
                    # Add to schedule
                    schedule[index].append(course)
                    studentTests = updateStudentTests(students, studentTests, index)
                    added = True
                ### CASE 2B: NO REPEATED STUDENTS DURING TIMESLOT
                else:
                    repeatStudents = checkRepeatedStudents(courseStudent, schedule, course, index)
                    # If no repeat students or conflicts, add to schedule
                    if not repeatStudents and not studentConflict:
                        schedule[index].append(course)
                        studentTests = updateStudentTests(students, studentTests, index)
                        added = True

            ### CASE 2: NO ACCEPTABLE SLOTS FOUND
            # Add one to index if no slot found
            index += 1

    ### PRINT THE RESULTS
    for index in range(len(schedule)):
        day = index // MAX_TESTS
        time = index % MAX_TESTS
        print(f'\n\nDAY {day + 1}: SLOT {time + 1}:')
        print(*sorted(schedule[index]))
    print(f'Number of slots used: {len(schedule)}')

    export_excel(schedule, 4, "final_schedule.xlsx", False, 10)