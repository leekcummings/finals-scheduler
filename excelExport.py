from openpyxl import Workbook
from openpyxl.styles import *

def exportExcel(schedule, ntimes, path, compact=False, maxRows=15):
    wb = Workbook()
    sheet = wb.active

    cols = 3
    if compact: # Only change # cols in compact mode
        for slot in schedule:
            need = (len(slot)//maxRows)*2+3
            if need > cols:
                cols = need
    
    # Border definitions
    thin = Side(border_style="thin")
    thick = Side(border_style="thick")

    # Style for colunmn labels
    header = NamedStyle(name='header')
    header.font = Font(bold=True, size=14)
    header.alignment = Alignment(horizontal='center', vertical='center')
    header.border = Border(bottom=thick)

    for c in range(1,cols+3):
        sheet.cell(row=1, column=c).style = header
    for c in sheet[1]:
        sheet.column_dimensions[c.column_letter].width = 20

    # Labels
    sheet['A1'] = 'Exam Time'
    for c in range(2, cols+1): # Loop for compact columns
        if c%2 == 0:
            sheet.cell(row=1, column=c).value='Course'
        else:
            sheet.cell(row=1, column=c).value='Room'

    # Date and time inputs to be referenced in-sheet later
    sheet.cell(row=1, column=cols+2).value = 'Dates'
    sheet.cell(row=1, column=cols+2).font = Font(bold=True)
    for n in range(1, len(schedule)//ntimes+2):
        sheet.cell(row=n+1,column=cols+2).value = f"Day {n}"
        sheet.cell(row=n+1, column=cols+2).number_format = numbers.FORMAT_TEXT
    
    tRow = len(schedule)//ntimes+4
    sheet.cell(row=tRow, column=cols+2).value = 'Times'
    sheet.cell(row=tRow, column=cols+2).style = header
    sheet.cell(row=tRow, column=cols+2).font = Font(bold=True)
    for n in range(1, ntimes+1):
        sheet.cell(row=tRow+n, column=cols+2).value = f"Slot {n}"
        sheet.cell(row=tRow+n, column=cols+2).number_format = numbers.FORMAT_TEXT
    
    # Helper variables for formatting and references
    dataCol = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[cols+1] #stupidest implementation possible but it works lmao
    courseBorder = Border(bottom=thin, left=thin, right=thin, top=thin)
    roomBorder = Border(bottom=thin, top=thin, left=thin, right=thick)
    darkColor = PatternFill("solid",fgColor="FFa5c9ef")
    lightColor = PatternFill("solid",fgColor="FFd2e4f7") # NOTE: Maybe change colors to be more 'simmons'

    curRow = 2
    for n in range(len(schedule)):
        if len(schedule[n]) != 0:
            # Reference time and date
            sheet.cell(row=curRow, column=1).value = f'=_xlfn.TEXTJOIN(CHAR(10), TRUE, {dataCol}{n//ntimes+2}, {dataCol}{n%ntimes+tRow+1})'
            sheet.cell(row=curRow, column=1).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            if n%2==0:
                sheet.cell(row=curRow, column=1).fill = darkColor
            else:
                sheet.cell(row=curRow, column=1).fill = lightColor

            # Loop helpers for compact mode
            startRow = curRow
            maxRow = curRow+maxRows-1 

            curCol = 2

            sorted = schedule[n].copy()
            sorted.sort()
            for course in sorted:
                top = False # Used for border styling later
                if compact and curRow > maxRow: # Only change column in compact mode
                    curRow = startRow
                    curCol += 2
                    top = True
                
                # Set course and style 
                sheet.cell(row=curRow,column=curCol).value = course
                sheet.cell(row=curRow,column=curCol).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=curRow,column=curCol).border = courseBorder
                sheet.cell(row=curRow,column=curCol+1).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=curRow,column=curCol+1).border = roomBorder
                if n%2==0:
                    sheet.cell(row=curRow,column=curCol).fill = darkColor
                    sheet.cell(row=curRow,column=curCol+1).fill = darkColor
                else:
                    sheet.cell(row=curRow,column=curCol).fill = lightColor
                    sheet.cell(row=curRow,column=curCol+1).fill = lightColor
                
                # Ensure thick borders in compact mode
                if top:
                    sheet.cell(row=curRow,column=curCol).border = Border(bottom=thin, left=thin, right=thin, top=thick)
                    sheet.cell(row=curRow,column=curCol+1).border = Border(bottom=thin, left=thin, right=thick, top=thick)
                
                curRow+=1
            
            # Create thick border beneath current section, merge section label column
            if curCol == 2:
                for c in range(1, 3):
                    sheet.cell(row=curRow-1,column=c).border = Border(bottom=thick, left=thin, right=thin, top=thin)
                sheet.cell(row=curRow-1, column=3).border = Border(bottom=thick, left=thin, right=thick, top=thin)
                sheet.merge_cells(f"A{startRow}:A{curRow-1}")
            else:
                for c in range(1, curCol):
                    if c > 1 and c%2==1:
                        sheet.cell(row=maxRow,column=c).border = Border(bottom=thick, left=thin, right=thick, top=thin)
                    else:
                        sheet.cell(row=maxRow,column=c).border = Border(bottom=thick, left=thin, right=thin, top=thin)
                if curRow == startRow+1:
                    sheet.cell(row=curRow-1,column=curCol).border = Border(bottom=thick, left=thin, right=thin, top=thick)
                    sheet.cell(row=curRow-1,column=curCol+1).border = Border(bottom=thick, left=thin, right=thick, top=thick)
                else:
                    sheet.cell(row=curRow-1,column=curCol).border = Border(bottom=thick, left=thin, right=thin, top=thin)
                    sheet.cell(row=curRow-1,column=curCol+1).border = Border(bottom=thick, left=thin, right=thick, top=thin)
                sheet.merge_cells(f"A{startRow}:A{maxRow}")
                curRow = maxRow + 1

    wb.save(filename=path)
